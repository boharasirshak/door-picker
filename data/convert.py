# A Python program to convert the excel data inside the /data folders into a easy to read and consistent sqlite database
# The main motivation to make this is that excels are hard and inefficent to read in python and
# the data was not consistent to read programatically. Sqlite was chosen because it is easy to use.

# To read and understand what this code is doing, kindly put any excel file from the /data/ directory alongside this program
# This program uses various ways to parse the information. It is not a perfect code and things *can* improve, but
# hey it works for this use case and is fairly quick considering you have to commit 20k+ columns one-by-one.

import string
import sqlite3
from openpyxl import load_workbook

conn = sqlite3.connect("db.db")
cur = conn.cursor()

res = cur.execute(
    """
CREATE TABLE IF NOT EXISTS `products` (
    `id` INTEGER PRIMARY KEY AUTOINCREMENT,
    `name` TEXT NOT NULL,
    `height` TEXT NOT NULL,
    `width` TEXT NOT NULL,
    `color` TEXT NOT NULL,
    `handle_type` TEXT NOT NULL,
    `profile_system` TEXT NOT NULL,
    `image_path` TEXT NOT NULL,
    `opening_scheme` TEXT NOT NULL
);
"""
)

if res:
    print("Products created successfully")

res = cur.execute(
    """
CREATE TABLE IF NOT EXISTS `features` (
    `id` INTEGER PRIMARY KEY AUTOINCREMENT,
    `external_id` INTEGER NOT NULL,
    `name` TEXT NOT NULL,
    `per_unit` INTEGER NOT NULL,
    `product_id` INTEGER,
    FOREIGN KEY (product_id) REFERENCES products(id)
);
"""
)

if res:
    print("Features created successfully")


conn.commit()


class CustomExcelReader:
    def __init__(self, path: str, sheets_prefix: str, debug: bool = False) -> None:
        self.path = path.split("/")[-1].replace(".xlsx", "")
        self.wb = load_workbook(filename=path, read_only=True)
        self.sheets = self.wb.sheetnames
        self.cols = string.ascii_uppercase
        self.sheets_prefix = sheets_prefix
        self.MAX_GAP = 20
        self.debug = debug

    def extract_features(self, values: tuple, product_id: int, i: int):
        gap = 3

        while True:
            if i + gap > len(values) - 1:
                break

            data = values[i + gap]
            name = data[1]
            idx = data[2]
            unit = data[3]

            if name is None or idx is None or unit is None:
                break

            if not isinstance(idx, int) or not isinstance(unit, int):
                break

            name = str(name).strip()

            if idx == 490546 or idx == "490546":
                idx = 794253

            # print(f"Inserting ({idx}) {name} - {unit}  into {product_id}")
            cur.execute(
                "INSERT INTO features (name, external_id, per_unit, product_id) VALUES (?, ?, ?, ?)",
                (name, idx, unit, product_id),
            )
            gap += 1

    def extract_products(self, sheet_name: str):
        if not sheet_name.startswith(self.sheets_prefix):
            return

        print(f"Extracting products of {sheet_name}...")
        sheet = self.wb[sheet_name]
        values = list(sheet.values)
        color = None
        handle_type = None

        for i, value in enumerate(values):
            height = value[0]
            width = value[1]
            opening_scheme = value[1]

            if height is not None and width is not None:
                opening_scheme = values[i + 1][1]
                width = str(width).replace("створки ", "")

                # only works for one digit comma seperated schemes x,x,x
                image_type = opening_scheme[:5]

                opening_scheme = opening_scheme.split()
                if len(opening_scheme) == 3:
                    opening_scheme = (
                        f"{opening_scheme[0]} ({opening_scheme[1]} {opening_scheme[2]})"
                    )
                else:
                    opening_scheme = f"{opening_scheme[0]} (1 комплект)"

                color_cell = values[i + 1][4]

                if color_cell is not None:
                    color = color_cell

                if len(values[i + 1]) > 5 and values[i + 1][5] is not None:
                    handle_type = values[i + 1][5]

                if handle_type == "2-ст ру":
                    handle_type = "2-ст руч"

                if self.debug:
                    print(
                        f"Inserting height: {height} & width: {width} & profile_system: {opening_scheme} & color: {color} & handle_type: {handle_type}"
                    )

                sheet_name = str(sheet_name).strip()
                height = str(height).strip()
                width = str(width).strip()
                opening_scheme = str(opening_scheme).strip()
                color = str(color).strip()

                if handle_type:
                    handle_type = str(handle_type).strip()

                cur.execute(
                    """
                    INSERT INTO products (
                        name,
                        height,
                        width,
                        color,
                        handle_type,
                        profile_system,
                        opening_scheme,
                        image_path
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        sheet_name,
                        height,
                        width,
                        color or "",
                        handle_type or "1-ст руч",
                        self.path,
                        opening_scheme,
                        image_type + ".png",
                    ),
                )
                conn.commit()
                product_id = cur.lastrowid
                self.extract_features(values, product_id, i)
                conn.commit()

    def read_file(self):
        sheets = [
            sheetname
            for sheetname in self.wb.sheetnames
            if sheetname.startswith(self.sheets_prefix)
        ]

        for sheet in sheets:
            self.extract_products(sheet)


CustomExcelReader("Alumark S70.xlsx", "Alumark").read_file()

CustomExcelReader("Alutech W62,W72.xlsx", "Alutech").read_file()

CustomExcelReader("Krauss KRWD64.xlsx", "Krauss").read_file()

CustomExcelReader("Татпроф ТПТ 65.xlsx", "Tatprof").read_file()
