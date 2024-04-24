import os
from sys import exit

import sqlite3
import warnings
import darkdetect

from typing import Tuple, Callable
from tkinter import messagebox, filedialog

import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage

from PIL import Image
from tksheet import Sheet
import customtkinter as ctk

warnings.filterwarnings("ignore")

WIDTH = 900
HEIGHT = 600

if not os.path.exists("data/db.db"):
    messagebox.showerror(
        "Ошибка",
        "база данных data.db не найдена в папке data/, пожалуйста, свяжитесь с администраторами!",
    )
    exit(1)


if not os.path.exists("results"):
    os.makedirs("results")


conn = sqlite3.connect("data/db.db")
cur = conn.cursor()
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class InputFrame(ctk.CTkFrame):
    def __init__(
        self,
        parent,
        handle_height_input: Callable,
        handle_width_input: Callable,
        handle_color_input: Callable,
        handle_handle_type_input: Callable,
        handle_profile_system_input: Callable,
        handle_multiplier_input: Callable,
        widths=[],
        heights=[],
    ):
        super().__init__(parent)
        self.top_row = ctk.CTkFrame(self)
        self.top_row.grid(row=0, column=0, sticky="nsew")
        self.grid_columnconfigure(0, weight=1)
        self.top_row.grid_columnconfigure((0, 1, 2, 3), weight=1)

        self.height_label = ctk.CTkLabel(self.top_row, text="Высота:")
        self.height_label.grid(row=0, column=0, padx=10, pady=10)
        self.height_entry = ctk.CTkComboBox(
            self.top_row,
            values=heights,
            command=lambda event: handle_height_input(event, self.height_entry),
            width=190,
        )
        self.height_entry.grid(row=0, column=1, padx=5, pady=5)

        self.width_label = ctk.CTkLabel(self.top_row, text="Ширина:")
        self.width_label.grid(row=0, column=2, padx=5, pady=5)
        self.width_entry = ctk.CTkComboBox(
            self.top_row,
            values=widths,
            command=lambda event: handle_width_input(event, self.width_entry),
            width=230,
        )
        self.width_entry.grid(row=0, column=3, padx=5, pady=5)

        # Bottom row for Color, Handle Type, and Profile System
        self.bottom_row = ctk.CTkFrame(self, width=WIDTH)
        self.bottom_row.grid(row=1, column=0, sticky="nsew")
        self.bottom_row.grid_columnconfigure((0, 1, 2, 3, 4), weight=1)

        self.color_label = ctk.CTkLabel(self.bottom_row, text="Цвет:")
        self.color_label.grid(row=0, column=0, padx=5, pady=5)
        self.color_dropdown = ctk.CTkComboBox(
            self.bottom_row,
            values=["белый", "чёрный", "серебро", "без окраса"],
            command=lambda event: handle_color_input(event, self.color_dropdown),
            width=120,
        )
        self.color_dropdown.grid(row=0, column=1, padx=10, pady=10, sticky="ew")

        self.handle_type_label = ctk.CTkLabel(self.bottom_row, text="Ручка:")
        self.handle_type_label.grid(row=0, column=2, padx=5, pady=5)
        self.handle_type_dropdown = ctk.CTkComboBox(
            self.bottom_row,
            values=["", "2-ст руч"],
            command=lambda event: handle_handle_type_input(
                event, self.handle_type_dropdown
            ),
            width=100,
        )
        self.handle_type_dropdown.grid(row=0, column=3, padx=5, pady=5, sticky="ew")

        self.profile_system_label = ctk.CTkLabel(
            self.bottom_row, text="Профильная система:"
        )
        self.profile_system_label.grid(row=0, column=4, padx=5, pady=5)
        self.profile_system_dropdown = ctk.CTkComboBox(
            self.bottom_row,
            values=[
                "Alumark S70",
                "Alutech W62,W72",
                "Krauss KRWD64",
                "Татпроф ТПТ 65",
            ],
            command=lambda event: handle_profile_system_input(
                event, self.profile_system_dropdown
            ),
        )
        self.profile_system_dropdown.grid(row=0, column=5, padx=5, pady=5, sticky="ew")

        self.multipler_label = ctk.CTkLabel(self.bottom_row, text="Количество:")
        self.multipler_label.grid(row=0, column=6, padx=5, pady=5)
        self.multiplier_dropdown = ctk.CTkComboBox(
            self.bottom_row,
            values=["1", "2", "3", "4", "5", "6", "7", "8", "9", "10"],
            command=lambda event: handle_multiplier_input(
                event, self.multiplier_dropdown
            ),
            width=60,
        )
        self.multiplier_dropdown.grid(row=0, column=7, padx=5, pady=5, sticky="ew")

        # Configure the main frame's row weights
        self.grid_rowconfigure(0, weight=1)  # Top row
        self.grid_rowconfigure(1, weight=1)  # Bottom row


class ImageFrame(ctk.CTkFrame):
    def __init__(self, master, size: Tuple[int, int] = (160, 160)):
        super().__init__(master)
        self.size = size

    def set_image(self, image: Image.Image):
        self.image = ctk.CTkImage(image, size=self.size)


class ImageRowFrame(ctk.CTkScrollableFrame):
    def __init__(
        self,
        parent,
        handle_click: Callable,
        images=[],
    ):
        super().__init__(parent)
        self.img_frames: list[ctk.CTkButton] = []
        self.update_images(images)
        self.handle_click = handle_click

    def update_images(self, image_paths: list[str]):
        for img_frame in self.img_frames:
            img_frame.destroy()
        self.img_frames.clear()

        number_of_images = len(image_paths)
        self.number_of_images = number_of_images
        images_per_row = 3
        number_of_rows = (number_of_images + images_per_row - 1) // images_per_row

        for row in range(number_of_rows):
            for col in range(images_per_row):
                self.grid_columnconfigure(col, weight=1)

        img_count = 0
        for row in range(number_of_rows):
            for col in range(images_per_row):
                if img_count >= number_of_images:
                    break
                image_path = image_paths[img_count]
                img_frame = ImageFrame(
                    self,
                )
                img = Image.open("data/imgs/" + image_path)
                img.thumbnail((160, 160), Image.Resampling.LANCZOS)
                img_frame.set_image(img)

                img_btn = ctk.CTkButton(
                    self,
                    image=img_frame.image,
                    bg_color="transparent",
                    fg_color="transparent",
                    hover_color=("gray75", "gray25"),
                    text=image_path,
                    text_color="gray25",
                )
                img_btn._command = lambda b=img_btn: self.handle_click(b)
                img_btn.grid(row=row, column=col, padx=10, pady=10, sticky="nsew")
                self.img_frames.append(img_btn)
                img_count += 1


class ExcelFrame(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent)

        # Widgets for Excel display frame (placeholder)
        # self.excel_data_label = ctk.CTkLabel(self, text="Data")
        # self.excel_data_label.grid(row=0, column=0, padx=10, pady=10)
        self.sheet = Sheet(
            self,
            header=["id", "name", "per unit", "Общий"],
            # show_x_scrollbar=False,
            zoom=150,
            data=[],
            width=WIDTH + 165,
            theme="dark" if darkdetect.isDark() else "light blue",
            default_column_width=110,
        )
        self.sheet.enable_bindings()
        self.sheet.grid(row=0, column=1, sticky="nswe", pady=10, padx=10)


class ButtonFrame(ctk.CTkFrame):
    def __init__(self, parent, save_data_callback):
        super().__init__(parent)
        self.save_data_callback = save_data_callback

        # Save Data Button
        self.save_button = ctk.CTkButton(
            self, text="Сохранить данные", command=self.on_save_data
        )
        self.save_button.grid(
            row=0, column=0, padx=10, pady=10, sticky="nsew", columnspan=2
        )

        self.grid_columnconfigure(0, weight=1)

    def on_generate_data(self):
        self.generate_data_callback()

    def on_save_data(self):
        self.save_data_callback()


class App(ctk.CTk):
    height_input: str = "высота 630мм-2400мм"
    width_input: str = "ширина до 3000мм"
    color_input: str = "белый"
    handle_type_input: str = ""
    profile_system_input: str = "Alumark S70"
    img_name: str = "*"
    multiplier_input: str = "1"
    all_data = []

    def __init__(self, fg_color: str | Tuple[str, str] | None = None, **kwargs):
        super().__init__(fg_color, **kwargs)

        self.title("Система определения конфигурации двери")
        self.geometry(f"{WIDTH}x{HEIGHT}")
        self.resizable(False, False)
        self.iconbitmap("data/logo.ico")

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        heights = cur.execute("SELECT DISTINCT height FROM products").fetchall() or []
        heights = [str(height[0]) for height in heights]
        widths = [
            "ширина до 3000мм",
            "ширина до 4000мм",
            "ширина до 5000мм",
            "ширина до 6000мм",
            # "ширина 3000 - 4000мм",
            # "ширина 4000 - 5000мм",
            # "ширина 5000 - 6000мм",
        ]

        self.input_frame = InputFrame(
            self,
            handle_height_input=self.handle_height_input,
            handle_width_input=self.handle_width_input,
            handle_color_input=self.handle_color_input,
            handle_handle_type_input=self.handle_handle_type_input,
            handle_profile_system_input=self.handle_profile_system_input,
            handle_multiplier_input=self.handle_multiplier_input,
            heights=heights,
            widths=widths,
        )
        self.input_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

        self.input_frame.grid_columnconfigure(0, weight=2)
        self.input_frame.grid_columnconfigure(3, weight=2)

        self.image_frame = ImageRowFrame(
            self, images=[], handle_click=self.handle_img_click
        )

        self.image_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")

        self.image_frame.grid_columnconfigure((0, 1, 2), weight=1)

        self.excel_frame = ExcelFrame(self)
        self.excel_frame.grid(
            row=2, column=0, padx=5, pady=5, sticky="nsew", columnspan=3
        )
        self.grid_rowconfigure(2, weight=3)

        self.button_frame = ButtonFrame(self, self.save_data)
        self.button_frame.grid(row=3, column=0, padx=10, pady=10, sticky="ew")

        self.button_frame.grid_columnconfigure((0, 1), weight=1)
        self.button_frame.save_button.grid_configure(sticky="ew")
        self._search_data()

    def handle_height_input(self, event, entry: ctk.CTkEntry):
        self.height_input = entry.get()
        self._search_data()

    def handle_width_input(self, event, entry: ctk.CTkEntry):
        self.width_input = entry.get()
        self._search_data()

    def handle_color_input(self, event, entry: ctk.CTkComboBox):
        self.color_input = entry.get()
        self._search_data()

    def handle_handle_type_input(self, event, entry: ctk.CTkComboBox):
        self.handle_type_input = entry.get()
        self._search_data()

    def handle_profile_system_input(self, event, entry: ctk.CTkComboBox):
        self.profile_system_input = entry.get()
        self._search_data()

    def handle_multiplier_input(self, event, entry: ctk.CTkComboBox):
        self.multiplier_input = entry.get()
        self._search_data()

    def handle_img_click(self, btn: ctk.CTkButton):
        self.img_name = btn.cget("text")

        # TODO: Fix the hover color
        for img_btn in self.image_frame.img_frames:
            if img_btn.cget("text") == self.img_name:
                img_btn.configure(bg_color="black")
            else:
                img_btn.configure(bg_color="transparent")

        self._search_data()

    def _search_data(self):
        if self.width_input == "ширина до 4000мм":
            width_input = ["ширина 3000 - 4000мм", "ширина до 4000мм"]

        elif self.width_input == "ширина до 5000мм":
            width_input = ["ширина 4000 - 5000мм", "ширина до 5000мм"]

        elif self.width_input == "ширина до 6000мм":
            width_input = ["ширина 5000 - 6000мм", "ширина до 6000мм"]

        else:
            width_input = [self.width_input]

        placeholders = ", ".join("?" for _ in width_input)

        res = cur.execute(
            f"""
            SELECT 
            f.external_id as id,
            f.name,
            f.per_unit, 
            p.opening_scheme,
            p.height,
            p.width,
            p.color,
            p.profile_system,
            p.image_path,
            p.id as product_id
            FROM products p
            LEFT JOIN features f ON f.product_id = p.id
            WHERE color = ? AND handle_type = ? AND profile_system = ? AND height = ? AND width IN ({placeholders})
            """,
            (
                self.color_input,
                self.handle_type_input,
                self.profile_system_input,
                self.height_input,
                *width_input,
            ),
        )
        all_data = res.fetchall()
        if not all_data:
            # print(
            #     self.color_input,
            #     self.handle_type_input,
            #     self.profile_system_input,
            #     self.height_input,
            #     width_input,
            # )
            all_data = []

        all_images = [data[8] for data in all_data]
        image_data = list(set(all_images))

        if self.img_name != "*":
            all_data = [data for data in all_data if data[8] == self.img_name]
            self.button_frame.save_button.configure(
                text=f"Сохранить данные - {self.img_name.replace('.png', '')}"
            )
        else:
            self.button_frame.save_button.configure(text="Сохранить данные")

        excel_data = [list(data[:3]) for data in all_data]

        multiplier = (
            1 if not self.multiplier_input.isdigit() else int(self.multiplier_input)
        )

        for data in excel_data:
            if str(data[2]).isdigit():
                data.append(int(data[2] * multiplier))
            else:
                data.append(data[2])

        self.image_frame.update_images(image_data)
        self.excel_frame.sheet.set_sheet_data(excel_data)

        self.all_data = all_data

    def generate_excel(self, entries: dict):
        multiplier = (
            1
            if not str(self.multiplier_input).isdigit()
            else int(self.multiplier_input)
        )

        for _, data in entries.items():
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = f"{data['color']} {data['handle_type']}"
                ws["A1"] = data["height"]
                ws["B1"] = data["width"]
                ws["B2"] = data["scheme"]
                ws["B4"] = "наименование"
                ws["C4"] = "артикул"
                ws["D4"] = "на ед"
                ws["E4"] = "Общий"

                ws["B4"].border = thin_border
                ws["C4"].border = thin_border
                ws["D4"].border = thin_border
                ws["E4"].border = thin_border

                for i, feature in enumerate(data["features"], start=5):
                    ws[f"B{i}"] = feature["name"]
                    ws[f"C{i}"] = feature["id"]
                    ws[f"D{i}"] = feature["per_unit"]
                    per_unit = str(feature["per_unit"])

                    if per_unit and per_unit.isdigit():
                        ws[f"E{i}"] = int(per_unit) * multiplier
                    else:
                        ws[f"E{i}"] = feature["per_unit"]

                    ws[f"B{i}"].border = thin_border
                    ws[f"C{i}"].border = thin_border
                    ws[f"D{i}"].border = thin_border
                    ws[f"E{i}"].border = thin_border

                img_path = os.path.join("data", "imgs", data["image_path"])
                if not os.path.exists(img_path):
                    messagebox.showerror(
                        "Изображение не найдено",
                        f"Изображение {img_path} не найдено. Пожалуйста, свяжитесь с администраторами!",
                    )
                    continue

                img = OpenpyxlImage(img_path)
                ws.add_image(img, "G4")

                fname = f"{data['profile']} - {data['color']} - {data['scheme']} - {data['height']} x {data['width']}.xlsx"

                path = filedialog.asksaveasfilename(
                    title="Сохраните файл Excel",
                    defaultextension=".xlsx",
                    confirmoverwrite=True,
                    filetypes=[("Файлы Excel", "*.xlsx")],
                    initialfile=fname,
                )

                wb.save(path)

                messagebox.showinfo(
                    "Успех",
                    f"Успешно сохраненный файл",
                )

            except PermissionError:
                messagebox.showerror(
                    "Ошибка разрешения",
                    "Пожалуйста, закройте файл Excel перед сохранением данных!",
                )

            except Exception as e:
                messagebox.showerror(
                    "Ошибка",
                    f"При сохранении данных произошла ошибка: {e}",
                )

    def save_data(self):
        entries = {}

        for data in self.all_data:
            product_id = data[-1]

            product = (
                cur.execute(
                    """
                SELECT 
                    height, 
                    width, 
                    color, 
                    handle_type, 
                    image_path, 
                    opening_scheme,
                    profile_system
                FROM products 
                WHERE id = ?
            """,
                    [int(product_id)],
                ).fetchone()
                or None
            )

            if product is None:
                messagebox.showerror(
                    "не найдено",
                    f"Товар с идентификатором {product_id} не найден в базе данных. Пожалуйста, свяжитесь с администраторами!",
                )
                continue

            height, width, color, handle_type, img_path, scheme, profile = product

            entries[product_id] = {
                "height": height,
                "width": width,
                "color": color,
                "handle_type": handle_type,
                "image_path": img_path,
                "scheme": scheme,
                "profile": profile,
                "features": [],
            }

            features = (
                cur.execute(
                    """
                SELECT
                    f.external_id as id,
                    f.name,
                    f.per_unit
                FROM products p
                LEFT JOIN features f ON f.product_id = p.id
                WHERE p.id = ?
            """,
                    [int(product_id)],
                ).fetchall()
                or []
            )

            if len(features) == 0:
                messagebox.showerror(
                    "Нечего спасать",
                    f"Продукт с идентификатором {product_id} не имеет характеристик в базе данных. Пожалуйста, свяжитесь с администраторами!",
                )
                continue

            for feature in features:
                entries[product_id]["features"].append(
                    {
                        "id": feature[0],
                        "name": feature[1],
                        "per_unit": feature[2],
                    }
                )

        self.generate_excel(entries)


if __name__ == "__main__":
    ctk.set_default_color_theme("dark-blue")
    app = App()
    app.mainloop()
